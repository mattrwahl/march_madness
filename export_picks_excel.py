"""
Export a full pick-by-pick breakdown for the v6-fixed-8-geomean model to Excel.
Model: coreB feature set (seed_rank_gap, conf_tourney_wins, dfi, tsi), geomean weights.
Output: data/v6_geomean_picks_analysis.xlsx
"""

import sys
import sqlite3

sys.path.insert(0, "C:/Users/user/sports_betting")
sys.path.insert(0, "C:/Users/user/sports_betting/march_madness")

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from processors.features import (
    load_eligible_teams,
    compute_composite_features,
    FEATURES_V6B,
    _build_feature_matrix_v6,
    select_picks_v6,
    V6_GEOMEAN_W,
)
from processors.model import simulate_pick_payout, ROUND_LABELS
from config import ALL_SEASONS, VAL_SEASONS, TEST_SEASONS, TRAIN_SEASONS

DB_PATH  = "C:/Users/user/sports_betting/march_madness/data/march_madness.db"
OUT_PATH = "C:/Users/user/sports_betting/march_madness/data/v6_geomean_picks_analysis.xlsx"


def _zscore_raw(values):
    """Z-score without negating (lower raw = lower z)."""
    mean = np.nanmean(values)
    std  = np.nanstd(values)
    if std > 0:
        return (values - mean) / std
    return values - mean


def build_rows(conn):
    rows = []
    for season in ALL_SEASONS:
        teams = load_eligible_teams(conn, season)
        if not teams:
            print("  [WARN] No eligible teams for season {}, skipping.".format(season))
            continue
        compute_composite_features(teams)
        X = _build_feature_matrix_v6(teams, FEATURES_V6B)
        picks = select_picks_v6(teams, V6_GEOMEAN_W, "coreB", 8)
        pick_ids = {p["team_id"] for p in picks}
        srg_vals = np.array([float(t.get("seed_rank_gap",     0) or 0) for t in teams])
        ctw_vals = np.array([float(t.get("conf_tourney_wins", 0) or 0) for t in teams])
        dfi_vals = np.array([float(t.get("dfi",               0) or 0) for t in teams])
        tsi_vals = np.array([float(t.get("tsi",               0) or 0) for t in teams])
        srg_z_all = _zscore_raw(srg_vals)
        ctw_z_all = _zscore_raw(ctw_vals)
        dfi_z_all = _zscore_raw(dfi_vals)
        tsi_z_all = _zscore_raw(tsi_vals)
        team_idx = {t["team_id"]: i for i, t in enumerate(teams)}
        if season in TRAIN_SEASONS:
            split = "Train"
        elif season in VAL_SEASONS:
            split = "Val"
        elif season in TEST_SEASONS:
            split = "Test"
        else:
            split = "Historical"
        for pick in picks:
            team_id   = pick["team_id"]
            pick_rank = pick["pick_rank"]
            other_ids = pick_ids - {team_id}
            tiered_bet = 37.50 if pick_rank <= 4 else 12.50
            flat_payout, round_exit_flat = simulate_pick_payout(
                conn, team_id, season,
                initial_bet=25.0,
                cash_out_round=8,
                other_pick_ids=other_ids,
                bet_style="flat",
            )
            tiered_payout, _ = simulate_pick_payout(
                conn, team_id, season,
                initial_bet=tiered_bet,
                cash_out_round=8,
                other_pick_ids=other_ids,
                bet_style="flat",
            )
            round_exit = round_exit_flat
            if round_exit is not None:
                round_exit_label = ROUND_LABELS.get(round_exit, "?")
            else:
                round_exit_label = "?"
            flat_units   = round((flat_payout   - 25.0)       / 100.0, 4)
            tiered_units = round((tiered_payout - tiered_bet) / 100.0, 4)
            idx = team_idx[team_id]
            row = {
                "Season":       season,
                "Split":        split,
                "Team":         pick.get("team_name", ""),
                "Seed":         pick.get("seed", ""),
                "Region":       pick.get("region", ""),
                "Pick_Rank":    pick_rank,
                "Model_Score":  round(float(pick.get("model_score", 0)), 4),
                "srg_raw":      round(float(pick.get("seed_rank_gap",     0) or 0), 4),
                "ctw_raw":      round(float(pick.get("conf_tourney_wins", 0) or 0), 4),
                "dfi_raw":      round(float(pick.get("dfi",               0) or 0), 4),
                "tsi_raw":      round(float(pick.get("tsi",               0) or 0), 4),
                "srg_z":        round(float(srg_z_all[idx]),  4),
                "ctw_z":        round(float(ctw_z_all[idx]),  4),
                "dfi_z":        round(float(dfi_z_all[idx]),  4),
                "tsi_z":        round(float(tsi_z_all[idx]),  4),
                "Result":       round_exit_label,
                "Flat_Bet":     25.0,
                "Flat_Units":   flat_units,
                "Tiered_Bet":   tiered_bet,
                "Tiered_Units": tiered_units,
            }
            rows.append(row)
        print("  Season {} ({}): {} picks processed.".format(season, split, len(picks)))
    rows.sort(key=lambda r: (r["Season"], r["Pick_Rank"]))
    return rows


COLUMNS = [
    "Season", "Split", "Team", "Seed", "Region",
    "Pick_Rank", "Model_Score",
    "srg_raw", "ctw_raw", "dfi_raw", "tsi_raw",
    "srg_z",   "ctw_z",   "dfi_z",   "tsi_z",
    "Result",
    "Flat_Bet", "Flat_Units",
    "Tiered_Bet", "Tiered_Units",
]

INTEGER_COLS = {"Season", "Seed", "Pick_Rank"}
STRING_COLS  = {"Split", "Team", "Region", "Result"}

FILL_TRAIN = PatternFill("solid", fgColor="FFFFE0")  # light yellow
FILL_VAL   = PatternFill("solid", fgColor="D6EAF8")  # light blue
FILL_TEST  = PatternFill("solid", fgColor="D5F5E3")  # light green
FILL_POS   = PatternFill("solid", fgColor="C6EFCE")  # green: positive units
FILL_NEG   = PatternFill("solid", fgColor="FFC7CE")  # red: negative units

SPLIT_FILLS = {"Train": FILL_TRAIN, "Val": FILL_VAL, "Test": FILL_TEST}

NUM_FMT = "0.00"


def write_excel(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Picks"

    # --- Header row ---
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # --- Data rows ---
    for row_idx, row in enumerate(rows, start=2):
        split    = row.get("Split", "")
        row_fill = SPLIT_FILLS.get(split)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_fill:
                cell.fill = row_fill
            if col_name not in INTEGER_COLS and col_name not in STRING_COLS:
                cell.number_format = NUM_FMT

    # --- Conditional formatting: Flat_Units and Tiered_Units ---
    flat_col_idx   = COLUMNS.index("Flat_Units")   + 1
    tiered_col_idx = COLUMNS.index("Tiered_Units") + 1
    last_data_row  = 1 + len(rows)

    for col_idx in (flat_col_idx, tiered_col_idx):
        col_letter = get_column_letter(col_idx)
        data_range = "{0}2:{0}{1}".format(col_letter, last_data_row)
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_POS),
        )
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=FILL_NEG),
        )

    # --- Auto-size columns ---
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for r in range(2, last_data_row + 1):
            cell_val = ws.cell(row=r, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = max_len + 2
    wb.save(path)


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    print("Building pick rows for all seasons...")
    rows = build_rows(conn)
    conn.close()

    print("")
    print("Writing Excel to: {}".format(OUT_PATH))
    write_excel(rows, OUT_PATH)

    print("")
    print("=" * 55)
    print("Total rows written: {}".format(len(rows)))
    print("")

    for split in ("Train", "Val", "Test", "Historical"):
        split_rows = [r for r in rows if r["Split"] == split]
        if not split_rows:
            continue
        flat_total   = sum(r["Flat_Units"]   for r in split_rows)
        tiered_total = sum(r["Tiered_Units"] for r in split_rows)
        seasons = len(set(r["Season"] for r in split_rows))
        print("  {:<10s}  seasons={}  picks={}  flat_units={:+.2f}  tiered_units={:+.2f}".format(
            split, seasons, len(split_rows), flat_total, tiered_total))

    print("")
    print("Output file: {}".format(OUT_PATH))
    print("Done.")


if __name__ == "__main__":
    main()
